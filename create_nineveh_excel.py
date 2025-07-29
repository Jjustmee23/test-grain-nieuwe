import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows

# Data based on the Nineveh industrial areas document
data = {
    'Nr.': list(range(1, 47)),  # Serial numbers 1-46
    'Naam van het gebied': [
        'Nineveh', 'Al-Qayyarah raffinaderij', 'Hammam al-Alil', 'Al-Qayyarah', 'Dar Al-Hamid',
        'Al-Muhalabiya', 'Al-Rashidiya', 'Al-Qoush', 'Al-Baaj', 'Rabia', 'Tel Kaif', 'Sinjar',
        'Siba Sheikh Khidr', 'Al-Mahlabiya', 'Al-Shalalat', 'Al-Ayman', 'Al-Hud', 'Al-Kharsa al-Janubiya',
        'Dar Al-Mahsour', 'Al-Badiyah', 'Tal Afar', 'Al-Murag', 'Al-Mawsil', 'Al-Bahar', 'Al-Khidr',
        'Umm Al-Rabeein', 'Al-Sharqat', 'Al-Dour', 'Al-Douris', 'Al-Khidr', 'Al-Hadher', 'Al-Rumaitha',
        'Al-Uwair', 'Ashab Al-Mustafa', 'Al-Tifl', 'Al-Hudhud', 'Al-Muhassin', 'Al-Ruwad', 'Al-Mansoor',
        'Al-Muhtasib', 'Al-Fatah', 'Al-Dinain', 'Al-Akhoin', 'Hatra', 'Al-Hamra', 'Badush'
    ],
    'Totale oppervlakte (m²)': [
        2059, 1544, 1544, 1544, 1287, 1287, 1287, 1029, 1029, 1029, 1029, 1029,
        1029, 1029, 1029, 1029, 1029, 1029, 1029, 1029, 772, 772, 772, 772, 772,
        722, 722, 722, 578, 578, 578, 578, 578, 578, 578, 578, 578, 578, 578, 578,
        578, 578, 578, 2205, 2606, 576
    ],
    'Beschikbare capaciteit (m²)': [
        400, 300, 300, 300, 250, 250, 250, 200, 200, 200, 200, 200,
        200, 200, 200, 200, 200, 200, 200, 200, 150, 150, 150, 150, 150,
        125, 125, 125, 100, 100, 100, 100, 100, 100, 100, 100, 100, 100, 100, 100,
        100, 100, 100, 200, 150, 50
    ]
}

# Create DataFrame
df = pd.DataFrame(data)
print(f"DataFrame shape: {df.shape}")

# Create Excel file with proper formatting
wb = Workbook()
ws = wb.active
ws.title = "Industriële Gebieden"

# Add title
ws.merge_cells('A1:D1')
title_cell = ws['A1']
title_cell.value = "Schema voor de toewijzing van industriële gebieden – Klasse 3 – voor het jaar 2025 volgens besluit nr. 3 van 23/3/2025 – Provincie Nineveh"
title_cell.font = Font(bold=True, size=14)
title_cell.alignment = Alignment(horizontal='center')

# Add summary section (top left)
ws.cell(row=3, column=1, value="291065")
ws.cell(row=4, column=1, value="229310")
ws.cell(row=5, column=1, value="101374")
ws.cell(row=6, column=1, value="621749")

ws.cell(row=3, column=2, value="سنجار")
ws.cell(row=4, column=2, value="تلعفر")
ws.cell(row=5, column=2, value="البعاج")

# Add summary section (top right)
ws.cell(row=3, column=3, value="عدد النفوس")
ws.cell(row=4, column=3, value="3993128")
ws.cell(row=5, column=3, value="36297.534")
ws.cell(row=6, column=3, value="45372")

ws.cell(row=3, column=4, value="الطحين")
ws.cell(row=4, column=4, value="الحبوب")

# Add headers starting from row 8
headers = ['Nr.', 'Naam van het gebied', 'Totale oppervlakte (m²)', 'Beschikbare capaciteit (m²)']
for col, header in enumerate(headers, 1):
    cell = ws.cell(row=8, column=col, value=header)
    cell.font = Font(bold=True)
    cell.alignment = Alignment(horizontal='center')
    cell.fill = PatternFill(start_color="D7E4BC", end_color="D7E4BC", fill_type="solid")

# Add data starting from row 9
for row_idx, row_data in enumerate(dataframe_to_rows(df, index=False, header=False), 9):
    for col_idx, value in enumerate(row_data, 1):
        cell = ws.cell(row=row_idx, column=col_idx, value=value)
        cell.alignment = Alignment(horizontal='center')

# Add total row
total_row = 9 + len(df)
ws.cell(row=total_row, column=1, value="Totaal beschikbaar")
ws.cell(row=total_row, column=2, value="")
ws.cell(row=total_row, column=3, value="45372")
ws.cell(row=total_row, column=4, value="")

# Add borders
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

for row in ws.iter_rows(min_row=8, max_row=total_row, min_col=1, max_col=4):
    for cell in row:
        cell.border = thin_border

# Set column widths
ws.column_dimensions['A'].width = 8   # Nr.
ws.column_dimensions['B'].width = 30  # Naam van het gebied
ws.column_dimensions['C'].width = 25  # Totale oppervlakte (m²)
ws.column_dimensions['D'].width = 25  # Beschikbare capaciteit (m²)

# Add footnotes
footnote_row = total_row + 3
ws.cell(row=footnote_row, column=1, value="Opmerkingen:")
ws.cell(row=footnote_row + 1, column=1, value="1. Volgens brief van het kabinet van de Eerste Minister nr. 526/46/8/11/1/8 d.d. 18/3/2008 is een oppervlakte van (2606) donum toegewezen aan de industriële zone.")
ws.cell(row=footnote_row + 2, column=1, value="2. Volgens brief van het Ministerie van Grondgebruik nr. 16602 d.d. 18/10/2009 is een oppervlakte van (3283) donum toegewezen.")
ws.cell(row=footnote_row + 3, column=1, value="3. Een oppervlakte van (576) donum is toegewezen voor Al-Hamdaniya en een gebied onder Sinjar. Voor het district Al-Baaj is (1153) donum beschikbaar gesteld.")

# Save the file
wb.save('nineveh_industrial_areas.xlsx')
print("Nineveh industrial areas Excel file created: nineveh_industrial_areas.xlsx")
print(f"Total rows: {len(df)}")
print("Columns: Nr., Naam van het gebied, Totale oppervlakte (m²), Beschikbare capaciteit (m²)") 