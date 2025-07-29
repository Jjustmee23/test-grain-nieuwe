import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows

# Data based on the Diyala document structure - 47 rows
mill_names = [
    'دينوي', 'محسن الدوري الفيصل', 'الموصل', 'سنجار', 'البعاج', 'الجزيرة', 'تلعفر',
    'الزبير', 'النجف', 'كربلاء', 'الحلة', 'الديوانية', 'الناصرية', 'العمارة',
    'الكوت', 'الزبير', 'البصرة', 'القرنة', 'المنصورية', 'الديالى', 'بعقوبة',
    'الخالص', 'بلدروز', 'المقدادية', 'خانقين', 'مندلي', 'الزبير', 'النجف',
    'كربلاء', 'الحلة', 'الديوانية', 'الناصرية', 'العمارة', 'الكوت', 'الزبير',
    'البصرة', 'القرنة', 'المنصورية', 'الديالى', 'بعقوبة', 'الخالص', 'بلدروز',
    'المقدادية', 'خانقين', 'مندلي', 'الزبير', 'النجف', 'كربلاء'
]

capacities = [
    400, 300, 250, 200, 150, 125, 100, 50, 400, 300, 250, 200, 150, 125,
    100, 50, 400, 300, 250, 200, 150, 125, 100, 50, 400, 300, 250, 200,
    150, 125, 100, 50, 400, 300, 250, 200, 150, 125, 100, 50, 400, 300,
    250, 200, 150, 125, 100
]

net_grains = [
    2059, 1544, 1287, 1029, 772, 722, 578, 2205, 2606, 1102, 576, 2059,
    1544, 1287, 1029, 772, 722, 578, 2205, 2606, 1102, 576, 2059, 1544,
    1287, 1029, 772, 722, 578, 2205, 2606, 1102, 576, 2059, 1544, 1287,
    1029, 772, 722, 578, 2205, 2606, 1102, 576, 2059, 1544, 1287
]

# Check lengths
print(f"Mill names length: {len(mill_names)}")
print(f"Capacities length: {len(capacities)}")
print(f"Net grains length: {len(net_grains)}")

# Ensure all arrays have the same length (47)
target_length = 47
mill_names = mill_names[:target_length]
capacities = capacities[:target_length]
net_grains = net_grains[:target_length]

# Create DataFrame
data = {
    'ت': list(range(1, target_length + 1)),  # Serial numbers 1-47
    'اسم المطحنة': mill_names,
    'الطاقة التصميمية': capacities,
    'صافي الحبوب': net_grains
}

df = pd.DataFrame(data)
print(f"DataFrame shape: {df.shape}")

# Create Excel file with proper formatting
wb = Workbook()
ws = wb.active
ws.title = "مخطط تجهيز الحبوب"

# Add title
ws.merge_cells('A1:D1')
title_cell = ws['A1']
title_cell.value = "مخطط تجهيز الحبوب للحصة 3 / 2025 والتي تبدا اعتبارا من 23/3/2025 محافظة ديدوى"
title_cell.font = Font(bold=True, size=14)
title_cell.alignment = Alignment(horizontal='center')

# Add summary section (top left)
ws.cell(row=3, column=1, value="291065")
ws.cell(row=4, column=1, value="229310")
ws.cell(row=5, column=1, value="101374")
ws.cell(row=6, column=1, value="621749")

ws.cell(row=3, column=2, value="سنجار")
ws.cell(row=4, column=2, value="تلعفر")
ws.cell(row=5, column=2, value="البعاج")

# Add summary section (top right)
ws.cell(row=3, column=3, value="عدد التقوس")
ws.cell(row=4, column=3, value="3993128")
ws.cell(row=5, column=3, value="36297.534")
ws.cell(row=6, column=3, value="45372")

ws.cell(row=3, column=4, value="الطحين")
ws.cell(row=4, column=4, value="الحبوب")

# Add headers starting from row 8
headers = ['ت', 'اسم المطحنة', 'الطاقة التصميمية', 'صافي الحبوب']
for col, header in enumerate(headers, 1):
    cell = ws.cell(row=8, column=col, value=header)
    cell.font = Font(bold=True)
    cell.alignment = Alignment(horizontal='center')
    cell.fill = PatternFill(start_color="D7E4BC", end_color="D7E4BC", fill_type="solid")

# Add data starting from row 9
for row_idx, row_data in enumerate(dataframe_to_rows(df, index=False, header=False), 9):
    for col_idx, value in enumerate(row_data, 1):
        cell = ws.cell(row=row_idx, column=col_idx, value=value)
        cell.alignment = Alignment(horizontal='center')

# Add borders
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

for row in ws.iter_rows(min_row=8, max_row=8+len(df), min_col=1, max_col=4):
    for cell in row:
        cell.border = thin_border

# Set column widths
ws.column_dimensions['A'].width = 8   # ت
ws.column_dimensions['B'].width = 25  # اسم المطحنة
ws.column_dimensions['C'].width = 20  # الطاقة التصميمية
ws.column_dimensions['D'].width = 20  # صافي الحبوب

# Add footnotes
footnote_row = 8 + len(df) + 2
ws.cell(row=footnote_row, column=1, value="ملاحظة:")
ws.cell(row=footnote_row + 1, column=1, value="1. تم تخصيص كمية (2606) من حبوب المسطحة بناءً على كتاب الأمانة العامة لمجلس الوزراء م . ت / 526/46/6/1/1/8 بتاريخ 2008/3/18")
ws.cell(row=footnote_row + 2, column=1, value="2. تم تخصيص كمية (3283) من حبوب المسطحة بناءً على هوامش الوزير بموجب الكتاب رقم 16602 بتاريخ 2009/10/18")
ws.cell(row=footnote_row + 3, column=1, value="3. تم تخصيص كمية (576) من حبوب المسطحة لمطحنة البعاج وهي نصف الكمية المخصصة لقضاء البعاج (1153 طن حبوب) مع عدد الأفراد")

# Save the file
wb.save('diyala_batch_import.xlsx')
print("Diyala style Excel file created: diyala_batch_import.xlsx")
print(f"Total rows: {len(df)}")
print("Columns: ت, اسم المطحنة, الطاقة التصميمية, صافي الحبوب") 